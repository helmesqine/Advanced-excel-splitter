from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
from tkinter import ttk, PhotoImage
from openpyxl.styles import Border, Side, Alignment, Font, GradientFill
import xlsxwriter
from openpyxl.reader.excel import load_workbook
import os
import glob
import traceback
import tkinter as tk
from datetime import datetime

root = Tk()
root.iconbitmap("img/EX.ico")
root.title('Splitter/Merger')
root.geometry("500x425")
root.resizable(0, 0)

frame = Frame(root, width=500, height=400)
frame_merge = Frame(root, width=500, height=400)
frame_home = Frame(root, width=500, height=400)
help_frame = Frame(root, width=500, height=400)
help_frame_merge = Frame(root, width=500, height=400)
merge_frame = Frame(root, width=500, height=400)
about_frame = Frame(root, width=500, height=400)
choose_sheet_frame = Frame(root, width=500, height=425)
set_format_frame = Frame(root, width=500, height=425)

choose_sheet_background = PhotoImage(file="img/choose sheet name.png")
label_choose_sheet = Label(choose_sheet_frame, image=choose_sheet_background)
label_choose_sheet.place(x=0, y=0)

set_format_background = PhotoImage(file="img/set format.png")
label_set_format = Label(set_format_frame, image=set_format_background)
label_set_format.place(x=0, y=0)

spliter_background = PhotoImage(file="img/background5.png")
label_spliter = Label(frame, image=spliter_background)
label_spliter.place(x=0, y=0)

home_bg = PhotoImage(file="img/home1.png")
label_home = Label(frame_home, image=home_bg)
label_home.place(x=0, y=0)

merge_bg = PhotoImage(file="img/merge_background.png")
label_merge = Label(merge_frame, image=merge_bg)

frame_home.pack(side=BOTTOM)

file = ''

Tips_split = '\n   How to use Excel splitter?\n\
           \n✅Click import and select the Excel file you want to split.\n\
            \n✅ Select the header of the column that you want to split. \n\
           \n✅Select the type of the Excel file.\n\
            \n✅ Make sure that there are no spaces in the data entered in the cells.\n\
            \n✅ Choose a directory output and click split.\n\
            \n\
            \n\
             \n✅click clear if you wanna reset.'

Tips_merge = '\n   How to use Excel Merger?\n\
           \n✅select the Directory of the files you want to merge.\n\
           \n✅Enter name of file output and do not include ".xlsx" or    ".xls".\n\
            \n✅Select the type of the Excel file out put.\n\
            \n✅select the Directory output.\n\
             \n✅click merge.\n\
            \n\
            \n\
             \n✅click clear if you wanna reset.'


def exit_programm():
    MsgBox = tk.messagebox.askquestion('EX Spliter Merger', 'Are you sure you want to exit the application',
                                       icon='warning')
    if MsgBox == 'yes':
        root.destroy()
    else:
        pass


root.protocol('WM_DELETE_WINDOW', exit_programm)


def about_function():
    global excel_spliter
    global excel_merge
    global excel_home
    global help_frame
    global about_text_merge

    frame_home.pack_forget()
    frame.pack_forget()
    help_frame.pack_forget()
    merge_frame.pack_forget()
    help_frame_merge.pack_forget()
    about_text_merge.pack_forget()
    about_frame.pack_forget()

    excel_home.place_forget()
    excel_spliter.place_forget()
    excel_merge.place_forget()

    excel_home = Button(root, bg='white', text='H' + "\u0332" + 'ome', padx=20, pady=4, highlightthickness=0, bd=0.5,
                        command=home_fuction)
    excel_home.place(x=0, y=0)
    excel_spliter = Button(root, bg='white', text='S' + "\u0332" + 'plitter', padx=20, pady=4, highlightthickness=0,
                           bd=0.5, command=expliter_fuction)
    excel_spliter.place(x=78, y=0)
    excel_merge = Button(root, bg='white', text='M' + "\u0332" + 'erge', padx=20, pady=4, highlightthickness=0, bd=0.5,
                         command=merge_fuction)
    excel_merge.place(x=160, y=0)

    about_text_merge = tk.Text(about_frame, height=24, width=62)
    about_text_merge.bind("<Key>", lambda a: "break")
    about_text_merge.insert("1.0", 'copy right text here')
    about_text_merge.tag_add("center", "1.0", "end")

    about_text_merge.place(x=0, y=10)
    about_frame.pack(side=BOTTOM)


def home_fuction():
    global excel_spliter
    global excel_merge
    global excel_home
    global help_frame
    global about_text_merge

    excel_home.place_forget()
    excel_merge.place_forget()
    excel_spliter.place_forget()
    frame.pack_forget()
    help_frame.pack_forget()
    merge_frame.pack_forget()
    help_frame_merge.pack_forget()
    about_text_merge.pack_forget()
    about_frame.pack_forget()

    excel_home = Button(root, bg='white', bd=0.5, text='H' + "\u0332" + 'ome', padx=20, pady=2, command=home_fuction)
    excel_home.place(x=0, y=0)
    excel_spliter = Button(root, bg='white', text='S' + "\u0332" + 'plitter', padx=20, pady=4, highlightthickness=0,
                           bd=0.5, command=expliter_fuction)
    excel_spliter.place(x=78, y=0)
    excel_merge = Button(root, bg='white', text='M' + "\u0332" + 'erge', padx=20, pady=4, highlightthickness=0, bd=0.5,
                         command=merge_fuction)
    excel_merge.place(x=160, y=0)

    label_spliter.place(x=0, y=0)
    frame_home.pack(side=BOTTOM)


def expliter_fuction():
    global excel_spliter
    global excel_merge
    global excel_home
    global help_frame
    global about_text_merge

    excel_home.place_forget()
    excel_merge.place_forget()
    excel_spliter.place_forget()
    frame_home.pack_forget()
    merge_frame.pack_forget()
    help_frame.pack_forget()
    help_frame_merge.pack_forget()
    about_text_merge.pack_forget()
    about_frame.pack_forget()

    excel_home = Button(root, bg='white', bd=0.5, text='H' + "\u0332" + 'ome', padx=20, pady=4, command=home_fuction,
                        highlightthickness=0)
    excel_home.place(x=0, y=0)
    excel_spliter = Button(root, bg='white', text='S' + "\u0332" + 'plitter', padx=20, pady=2,
                           bd=0.5, command=expliter_fuction)
    excel_spliter.place(x=78, y=0)
    excel_merge = Button(root, bg='white', text='M' + "\u0332" + 'erge', padx=20, pady=4, highlightthickness=0, bd=0.5,
                         command=merge_fuction)
    excel_merge.place(x=160, y=0)

    frame.pack(side=BOTTOM)


def return_to_main():
    help_frame.pack_forget()
    help_text.pack_forget()
    frame.pack(side=BOTTOM)


def return_to_main_merge():
    help_frame_merge.pack_forget()
    help_text_merge.pack_forget()
    merge_frame.pack(side=BOTTOM)


def merge_fuction():
    global excel_spliter
    global excel_merge
    global excel_home
    global help_frame
    global about_text_merge

    excel_home.place_forget()
    excel_merge.place_forget()
    excel_spliter.place_forget()
    frame.pack_forget()
    frame_home.pack_forget()
    help_frame.pack_forget()
    help_frame_merge.pack_forget()
    about_text_merge.pack_forget()
    about_frame.pack_forget()

    excel_home = Button(root, bg='white', bd=0.5, text='H' + "\u0332" + 'ome', padx=20, pady=4, command=home_fuction,
                        highlightthickness=0)
    excel_home.place(x=0, y=0)
    excel_spliter = Button(root, bg='white', text='S' + "\u0332" + 'plitter', padx=20, pady=4, highlightthickness=0,
                           bd=0.5, command=expliter_fuction)
    excel_spliter.place(x=78, y=0)
    excel_merge = Button(root, bg='white', text='M' + "\u0332" + 'erge', padx=20, pady=2, bd=0.5,
                         command=merge_fuction)
    excel_merge.place(x=160, y=0)

    frame_home.pack_forget()
    label_merge.place(x=0, y=0)
    merge_frame.pack(side=BOTTOM)


def help_frame_fuction():
    frame.pack_forget()
    help_frame.pack(side=BOTTOM)
    help_text.pack(side=BOTTOM)
    return_to_main_button = Button(help_frame, bg='white', text='Back!', padx=40, pady=6, command=return_to_main)
    return_to_main_button.place(x=170, y=350)


def help_frame_merge_fuction():
    merge_frame.pack_forget()
    help_frame_merge.pack(side=BOTTOM)
    help_text_merge.pack(side=BOTTOM)
    return_to_main_button_merge = Button(help_frame_merge, bg='white', text='Back!', padx=40, pady=6,
                                         command=return_to_main_merge)
    return_to_main_button_merge.place(x=170, y=350)


def UploadAction():
    global file_read
    global save_file
    global selected_filename
    global combo_sheet
    save_file_not.place_forget()
    filename = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xls")])
    if filename == '':
        messagebox.showwarning('EX Spliter Merger', 'Choose a File!!')
        UploadAction()
    try:
        save_file.place_forget()
    except:
        pass
    try:
        combo_column_first.place_forget()
    except:
        pass
    try:
        combo_column_second.place_forget()
    except:
        pass
    try:
        combo_column_third.place_forget()
    except:
        pass
    save_file = Button(frame, bg='white', text='Split', padx=15, pady=10, command=get_file)
    save_file.place(x=160, y=70)
    split_filename = str(filename).split('/')
    selected_filename = split_filename[-1]
    choose_sheet_frame.pack()

    sheets = pd.ExcelFile(filename)
    names_sheets = sheets.sheet_names
    combo_sheet = ttk.Combobox(choose_sheet_frame, state="readonly", value=[*names_sheets])
    combo_sheet.place(x=195, y=220)
    combo_sheet.bind("<<ComboboxSelected>>", lambda event: after_select())

    frame.forget()
    excel_home.place_forget()
    about_button.place_forget()
    excel_spliter.place_forget()
    excel_merge.place_forget()
    choose_sheet_frame.pack()

    file_read = filename
    after_select()


def sheet_get_fuction():
    return sheet_get


def after_select():
    global sheet_get
    global root_select_quit_button
    sheet_get_first = str(combo_sheet.get())
    sheet_get = sheet_get_first
    root_select_quit_button = Button(choose_sheet_frame, bg='white', text='Ok', padx=15, pady=6,
                                     command=after_upload_action)
    root_select_quit_button.place(x=220, y=260)


def after_upload_action():
    return_file()
    sheet_get_fuction()
    global Browse_button
    global file, root
    global drop_column
    global drop_typefile
    global directory
    global combo_column
    global combo_typefile
    global text_box
    global filexls
    global filexlsx
    global split_button
    global clear_button
    global selected_filename
    global filename
    global sheet_get
    global re_headers

    choose_sheet_frame.pack_forget()

    frame.pack(side=BOTTOM)
    excel_home.place(x=0, y=0)
    excel_spliter.place(x=80, y=0)
    about_button.place(x=240, y=0)
    excel_merge.place(x=160, y=0)

    read_data()


def return_file():
    global file_read
    return file_read


def read_data():
    sheet_get_fuction()
    global file_read
    global combo_column_first
    global combo_column_second
    global combo_column_third

    save_file_not.place_forget()
    combo_column_first_not.place_forget()
    combo_column_second_not.place_forget()
    combo_column_third_not.place_forget()

    columns = pd.read_excel(file_read, sheet_get).columns

    combo_column_first = ttk.Combobox(frame, state="readonly", value=[*columns])
    combo_column_first.current(0)
    combo_column_first.place(x=80, y=140)

    combo_column_second = ttk.Combobox(frame, state="readonly", value=['', *columns])
    combo_column_second.place(x=80, y=170)

    combo_column_third = ttk.Combobox(frame, state="readonly", value=['', *columns])
    combo_column_third.place(x=80, y=200)


def get_file():
    global file_read
    global combo_column_first
    global combo_column_second
    global combo_column_third
    global unique_values_1
    global unique_values_2
    global unique_values_3
    global directory
    global desktop_directory

    global head_color
    global head_size
    global cells_size
    global cells_color
    global align
    global list_colors_codes

    list_colors = ['White', 'Black', 'Grey', 'Yellow', 'Red', 'Blue', 'Green', 'Brown', 'Pink', 'Orange',
                   'Purple']
    list_sizes = [8, 10, 11, 12, 14, 16, 18, 20, 22, 24, 26, 28, 36, 48, 72]
    list_colors_codes = {'White': ['FFFFFF'], 'Black': ['000000'], 'Grey': ['808080'], 'Yellow': ['FFFF00'],
                         'Red': ['FF0000'],
                         'Blue': ['0000FF'], 'Green': ['008000'], 'Brown': ['A52A2A'], 'Pink': ['FFC0CB'],
                         'Orange': ['FFA500'],
                         'Purple': ['800080']}
    ask_format = tk.messagebox.askquestion('EX Spliter Merger', 'set format to output files ??')
    if ask_format == 'yes':
        head_color = ttk.Combobox(set_format_frame, state="readonly", value=[*list_colors])
        head_color.current(3)
        head_color.place(x=60, y=100)

        head_size = ttk.Combobox(set_format_frame, value=[*list_sizes])
        head_size.current(2)
        head_size.place(x=60, y=140)

        cells_size = ttk.Combobox(set_format_frame, value=[*list_sizes])
        cells_size.current(2)
        cells_size.place(x=280, y=140)

        cells_color = ttk.Combobox(set_format_frame, state="readonly", value=[*list_colors])
        cells_color.current(0)
        cells_color.place(x=280, y=100)

        align = ttk.Combobox(set_format_frame, state="readonly", value=['right', 'center', 'left'])
        align.current(1)
        align.place(x=60, y=240)

        ok_button = Button(set_format_frame, bg='white', text='Ok', padx=15, pady=6,
                           command=return_to_page)
        ok_button.place(x=60, y=320)

        frame.forget()
        excel_home.place_forget()
        about_button.place_forget()
        excel_spliter.place_forget()
        excel_merge.place_forget()
        choose_sheet_frame.pack()
        choose_sheet_frame.pack_forget()

        set_format_frame.pack()
    else:
        head_color = ttk.Combobox(set_format_frame, state="readonly", value=[*list_colors])
        head_color.current(3)

        head_size = ttk.Combobox(set_format_frame, value=[*list_sizes])
        head_size.current(2)

        cells_size = ttk.Combobox(set_format_frame, value=[*list_sizes])
        cells_size.current(2)

        cells_color = ttk.Combobox(set_format_frame, state="readonly", value=[*list_colors])
        cells_color.current(0)

        align = ttk.Combobox(set_format_frame, state="readonly", value=['right', 'center', 'left'])
        align.current(1)

        after_set_format()


def return_to_page():
    set_format_frame.pack_forget()

    frame.pack(side=BOTTOM)
    excel_home.place(x=0, y=0)
    excel_spliter.place(x=80, y=0)
    about_button.place(x=240, y=0)
    excel_merge.place(x=160, y=0)
    after_set_format()


def after_set_format():
    global desktop_directory
    global directory
    try:
        if directory == '':
            messagebox.showwarning('EX Spliter Merger', 'Choose directory out put!!')
            choose_directory()
    except:
        pass
    desktop_directory = os.path.join(os.path.join(os.environ['USERPROFILE']),
                                     'Desktop') + '/' + 'EX Splitter Merger Stats'
    if not os.path.exists(desktop_directory):
        os.mkdir(desktop_directory)
    else:
        pass

    if str(combo_column_third.get()) == '':
        if str(combo_column_second.get()) == '':
            one_columns()
        else:
            two_columns()
    else:
        if str(combo_column_second.get()) == '':
            messagebox.showwarning('EX Spliter Merger', 'Choose column!!')
            return_file()
        else:
            three_columns()


def three_columns():
    global file_read
    global combo_column_first
    global combo_column_second
    global combo_column_third
    global unique_values_1
    global unique_values_2
    global unique_values_3
    global directory
    global desktop_directory
    global selected_filename

    data = pd.read_excel(file_read)

    excel_stats_name = desktop_directory + '/' + 'statistics_' + selected_filename
    excel_stats = xlsxwriter.Workbook(excel_stats_name)
    excel_stats_sheet = excel_stats.add_worksheet()

    excel_stats_sheet.write(0, 0, 'name')
    excel_stats_sheet.write(0, 1, 'count')
    row = 1

    unique_values_1 = data[combo_column_first.get()].unique()
    unique_values_2 = data[combo_column_second.get()].unique()
    unique_values_3 = data[combo_column_third.get()].unique()

    if str(unique_values_1) == str(unique_values_2) == str(unique_values_3):
        messagebox.showwarning('EX Spliter Merger', 'Choose different column!!')
        read_data()
    elif str(unique_values_1) == str(unique_values_2):
        messagebox.showwarning('EX Spliter Merger', 'Choose different column!!')
        read_data()
    elif str(unique_values_1) == str(unique_values_3):
        messagebox.showwarning('EX Spliter Merger', 'Choose different column!!')
        read_data()
    elif str(unique_values_2) == str(unique_values_3):
        messagebox.showwarning('EX Spliter Merger', 'Choose different columns!!')
        read_data()
    else:
        for unique_1 in unique_values_1:
            if '.00' in str(unique_1):
                unique_1 = str(pd.to_datetime(unique_1).date())
            elif ':00' in str(unique_1):
                unique_1 = str(pd.to_datetime(unique_1).date())
            first_data = data[data[combo_column_first.get()].astype('str').str.match("^" + str(unique_1) + "$") == True]
            for unique_2 in unique_values_2:
                if '.00' in str(unique_2):
                    unique_2 = str(pd.to_datetime(unique_2).date())
                elif ':00' in str(unique_2):
                    unique_2 = str(pd.to_datetime(unique_2).date())
                second_data = first_data[
                    first_data[combo_column_second.get()].astype('str').str.match("^" + str(unique_2) + "$") == True]
                for unique_3 in unique_values_3:
                    if '.00' in str(unique_3):
                        unique_3 = str(pd.to_datetime(unique_3).date())
                    elif ':00' in str(unique_3):
                        unique_3 = str(pd.to_datetime(unique_3).date())
                    third_data = second_data[
                        second_data[combo_column_third.get()].astype('str').str.match(
                            "^" + str(unique_3) + "$") == True]
                    if 'Empty DataFrame' in str(third_data):
                        print('Empty')
                    else:
                        input_excel = directory + '/' + str(unique_1) + '_' + str(unique_2) + '_' + str(
                            unique_3) + '.xlsx'
                        third_data.to_excel(input_excel, index=False)
                        dimensions = third_data.shape
                        excel_stats_sheet.write(row, 0, str(unique_1) + '_' + str(unique_2) + '_' + str(unique_3))
                        excel_stats_sheet.write(row, 1, str(dimensions[0]))
                        row += 1

                        head_color_code = list_colors_codes[head_color.get()][0]
                        cells_color_code = list_colors_codes[cells_color.get()][0]

                        print(head_color_code)
                        print(cells_color_code)

                        wb = load_workbook(input_excel)
                        ws = wb.active
                        font_cell = Font(size=cells_size.get(), bold=False)
                        head_font = Font(size=head_size.get(), bold=True)
                        border_cell = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                             bottom=Side(style='thin'))
                        fill_head = GradientFill(stop=(head_color_code, head_color_code))
                        fill_cells = GradientFill(stop=(cells_color_code, cells_color_code))
                        alignment = Alignment(horizontal=align.get())
                        for col in ws.rows:
                            for cell in col:
                                ws[cell.coordinate].font = font_cell
                                ws[cell.coordinate].border = border_cell
                                ws[cell.coordinate].alignment = alignment
                                ws[cell.coordinate].fill = fill_cells
                        max_value = 0
                        headers = [cell for cell in ws[1]]
                        for head in headers:
                            letter = str(head.coordinate).replace('1', '')
                            for cell in ws[letter]:
                                value = len(str(cell.value))
                                if value > max_value:
                                    max_value = value
                            ws.column_dimensions[letter].width = max_value
                            ws[head.coordinate].font = head_font
                            ws[head.coordinate].border = border_cell
                            ws[head.coordinate].fill = fill_head
                            max_value = 0
                        wb.save(input_excel)

    excel_stats.close()
    directory = ''


def two_columns():
    global file_read
    global combo_column_first
    global combo_column_second
    global combo_column_third
    global unique_values_1
    global unique_values_2
    global unique_values_3
    global directory
    global desktop_directory

    data = pd.read_excel(file_read)

    excel_stats_name = desktop_directory + '/' + 'statistics_' + selected_filename
    excel_stats = xlsxwriter.Workbook(excel_stats_name)
    excel_stats_sheet = excel_stats.add_worksheet()

    excel_stats_sheet.write(0, 0, 'name')
    excel_stats_sheet.write(0, 1, 'count')
    row = 1

    unique_values_1 = data[combo_column_first.get()].unique()
    unique_values_2 = data[combo_column_second.get()].unique()

    if str(unique_values_1) == str(unique_values_2):
        messagebox.showwarning('EX Spliter Merger', 'Choose different columns!!')
        read_data()
    else:
        for unique_1 in unique_values_1:
            if '.00' in str(unique_1):
                unique_1 = str(pd.to_datetime(unique_1).date())
            elif ':00' in str(unique_1):
                unique_1 = str(pd.to_datetime(unique_1).date())
            first_data = data[data[combo_column_first.get()].astype('str').str.match("^" + str(unique_1) + "$") == True]
            for unique_2 in unique_values_2:
                if '.00' in str(unique_2):
                    unique_2 = str(pd.to_datetime(unique_2).date())
                elif ':00' in str(unique_2):
                    unique_2 = str(pd.to_datetime(unique_2).date())
                second_data = first_data[
                    first_data[combo_column_second.get()].astype('str').str.match("^" + str(unique_2) + "$") == True]
                if 'Empty DataFrame' in str(second_data):
                    print('Empty')
                else:
                    input_excel = directory + '/' + str(unique_1) + '_' + str(unique_2) + '.xlsx'
                    second_data.to_excel(input_excel, index=False)
                    dimensions = second_data.shape
                    excel_stats_sheet.write(row, 0, str(unique_1) + '_' + str(unique_2))
                    excel_stats_sheet.write(row, 1, str(dimensions[0]))
                    row += 1

                    head_color_code = list_colors_codes[head_color.get()][0]
                    cells_color_code = list_colors_codes[cells_color.get()][0]

                    wb = load_workbook(input_excel)
                    ws = wb.active
                    font_cell = Font(size=cells_size.get(), bold=False)
                    head_font = Font(size=head_size.get(), bold=True)
                    border_cell = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    fill_head = GradientFill(stop=(head_color_code, head_color_code))
                    fill_cells = GradientFill(stop=(cells_color_code, cells_color_code))
                    alignment = Alignment(horizontal=align.get())
                    for col in ws.rows:
                        for cell in col:
                            ws[cell.coordinate].font = font_cell
                            ws[cell.coordinate].border = border_cell
                            ws[cell.coordinate].alignment = alignment
                            ws[cell.coordinate].fill = fill_cells
                    max_value = 0
                    headers = [cell for cell in ws[1]]
                    for head in headers:
                        letter = str(head.coordinate).replace('1', '')
                        for cell in ws[letter]:
                            value = len(str(cell.value))
                            if value > max_value:
                                max_value = value
                        ws.column_dimensions[letter].width = max_value
                        ws[head.coordinate].font = head_font
                        ws[head.coordinate].border = border_cell
                        ws[head.coordinate].fill = fill_head
                        max_value = 0
                    wb.save(input_excel)
    excel_stats.close()
    directory = ''


def choose_directory():
    global Browse_button
    global file, root
    global drop_column
    global drop_typefile
    global directory
    global combo_column
    global combo_typefile
    global text_box
    global filexls
    global filexlsx
    global split_button
    global clear_button
    directory = filedialog.askdirectory()
    if len(directory) == 0:
        messagebox.showwarning('EX Spliter Merger', 'Choose directory out put!!')
        choose_directory()
    Browse_button_pop.place_forget()
    Browse_button = Button(frame, text='Folder is Selected', command=choose_directory, bg='white')
    Browse_button.place(x=32, y=250)


def one_columns():
    global file_read
    global combo_column_first
    global combo_column_second
    global combo_column_third
    global unique_values_1
    global unique_values_2
    global unique_values_3
    global directory
    global desktop_directory

    data = pd.read_excel(file_read)

    excel_stats_name = desktop_directory + '/' + 'statistics_' + selected_filename
    excel_stats = xlsxwriter.Workbook(excel_stats_name)
    excel_stats_sheet = excel_stats.add_worksheet()

    excel_stats_sheet.write(0, 0, 'name')
    excel_stats_sheet.write(0, 1, 'count')
    row = 1

    unique_values_1 = data[combo_column_first.get()].unique()

    if str(unique_values_1) == '':
        messagebox.showwarning('EX Spliter Merger', 'Choose different columns!!')
        read_data()
    else:
        for unique_1 in unique_values_1:
            if '.00' in str(unique_1):
                unique_1 = str(pd.to_datetime(unique_1).date())
            elif ':00' in str(unique_1):
                unique_1 = str(pd.to_datetime(unique_1).date())
            first_data = data[data[combo_column_first.get()].astype('str').str.match("^" + str(unique_1) + "$") == True]
            if 'Empty DataFrame' in str(first_data):
                print('Empty')
            else:
                input_excel = directory + '/' + str(unique_1) + '.xlsx'
                first_data.to_excel(input_excel, index=False)
                dimensions = first_data.shape
                excel_stats_sheet.write(row, 0, str(unique_1))
                excel_stats_sheet.write(row, 1, str(dimensions[0]))
                row += 1

                head_color_code = list_colors_codes[head_color.get()][0]
                cells_color_code = list_colors_codes[cells_color.get()][0]

                wb = load_workbook(input_excel)
                ws = wb.active
                ws.print_options.verticalCentered  = True
                font_cell = Font(size=cells_size.get(), bold=False)
                head_font = Font(size=head_size.get(), bold=True)
                border_cell = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                fill_head = GradientFill(stop=(head_color_code, head_color_code))
                fill_cells = GradientFill(stop=(cells_color_code, cells_color_code))
                alignment = Alignment(horizontal=align.get())
                for col in ws.rows:
                    for cell in col:
                        ws[cell.coordinate].font = font_cell
                        ws[cell.coordinate].border = border_cell
                        ws[cell.coordinate].alignment = alignment
                        ws[cell.coordinate].fill = fill_cells
                max_value = 0
                headers = [cell for cell in ws[1]]
                for head in headers:
                    letter = str(head.coordinate).replace('1', '')
                    for cell in ws[letter]:
                        value = len(str(cell.value))
                        if value > max_value:
                            max_value = value
                    ws.column_dimensions[letter].width = max_value
                    ws[head.coordinate].font = head_font
                    ws[head.coordinate].border = border_cell
                    ws[head.coordinate].fill = fill_head
                    max_value = 0
                wb.save(input_excel)
    excel_stats.close()
    directory = ''


def clear_all():
    global save_file
    global directory
    directory = ''
    try:
        save_file.place_forget()
    except:
        pass
    try:
        combo_column_first.place_forget()
    except:
        pass
    try:
        combo_column_second.place_forget()
    except:
        pass
    try:
        combo_column_third.place_forget()
    except:
        pass
    try:
        Browse_button.place_forget()
    except:
        pass
    try:
        save_file_not.place(x=160, y=70)
    except:
        pass
    try:
        Browse_button_pop.place(x=32, y=250)
    except:
        pass
    try:
        combo_column_first_not.place(x=80, y=140)
    except:
        pass
    try:
        combo_column_second_not.place(x=80, y=170)
    except:
        pass
    try:
        combo_column_third_not.place(x=80, y=200)
    except:
        pass


def clear_button_merge():
    global output_name_entry
    global directory_choosing_merge_output
    global directory_choosing_merge
    output_name_entry.delete(0, END)
    text_box_2.delete(1.0, END)
    directory_choosing_merge_output = ''
    directory_choosing_merge = ''


def merge_files_excels():
    global directory_choosing_merge
    global directory_choosing_merge_output
    print(directory_choosing_merge_output + output_name_entry.get())
    if directory_choosing_merge == 0:
        messagebox.showwarning('EX Spliter Merger', 'Choose directory!!')
        upload_action_merge_function()
    if directory_choosing_merge_output == 0:
        messagebox.showwarning('EX Spliter Merger', 'Choose output directory!!')
        directory_choosing_merge_output_function()
    merge_directory_xlsx = glob.glob(directory_choosing_merge + '*.xlsx')
    merge_directory_xls = glob.glob(directory_choosing_merge + '*.xls')
    merge_general_directory = merge_directory_xlsx + merge_directory_xls
    excels = [pd.ExcelFile(merge_name) for merge_name in merge_general_directory]
    frames = [x.parse(x.sheet_names[0], header=None, index_col=None) for x in excels]
    frames[1:] = [df[1:] for df in frames[1:]]
    combined = pd.concat(frames)
    if 'xlsx' in str(combo_typefile_pop_merge.get()):
        combined.to_excel(directory_choosing_merge_output + output_name_entry.get() + '.xlsx', header=False,
                          index=False)
    elif '2003' in str(combo_typefile_pop_merge.get()):
        combined.to_excel(directory_choosing_merge_output + output_name_entry.get() + '.xls', header=False, index=False)


def upload_action_merge_function():
    global directory_choosing_merge
    directory_choosing_merge = filedialog.askdirectory()
    if len(directory_choosing_merge) == 0:
        messagebox.showwarning('EX Spliter Merger', 'Choose directory!!')
        upload_action_merge_function()
    directory_choosing_merge += '/'
    text_box_2.insert(INSERT, '\nSelected Directory is : ' + directory_choosing_merge)
    text_box_2.yview(END)


def directory_choosing_merge_output_function():
    global directory_choosing_merge_output
    directory_choosing_merge_output = filedialog.askdirectory()
    if len(directory_choosing_merge_output) == 0:
        messagebox.showwarning('EX Spliter Merger', 'Choose output directory!!')
        directory_choosing_merge_output_function()
    directory_choosing_merge_output += '/'
    text_box_2.insert(INSERT, 'Selected Output Directory is : ' + directory_choosing_merge_output)
    text_box_2.yview(END)


def show_error(self, *args):
    error = traceback.format_exception(*args)
    messagebox.showerror('EX Spliter Merger', 'something went wrong!!')
    with open('log.txt', 'a') as file_txt:
        today = datetime.now()
        file_txt.write("\nEX Spliter Merger " + str(today) + "\n")
        file_txt.writelines(error)
        file_txt.close()


combo_column_first_not = ttk.Combobox(frame, state="readonly", value=[])
combo_column_first_not.place(x=80, y=140)

combo_column_second_not = ttk.Combobox(frame, state="readonly", value=[])
combo_column_second_not.place(x=80, y=170)

combo_column_third_not = ttk.Combobox(frame, state="readonly", value=[])
combo_column_third_not.place(x=80, y=200)

import_button = Button(frame, bg='white', text='Import', padx=10, pady=10, command=UploadAction)
import_button.place(x=32, y=70)
Browse_button_pop = Button(frame, text='Select Folder', bg='white', command=choose_directory)
save_file_not = Button(frame, bg='white', text='No file', padx=15, pady=10, command='')
clear_button_pop = Button(frame, bg='white', text='Clear', padx=40, pady=6, command=lambda: clear_all())
showtips = Button(frame, bg='white', text='Help!', padx=40, pady=6, command=help_frame_fuction)
clear_button_pop.place(x=32, y=290)
showtips.place(x=32, y=330)
save_file_not.place(x=160, y=70)
Browse_button_pop.place(x=32, y=250)

about_text_merge = tk.Text(about_frame)
about_text_merge.tag_configure("center", justify='left')
about_button = Button(root, bg='white', bd=0.5, text='A' + "\u0332" + 'bout', padx=20, pady=4, highlightthickness=0,
                      command=about_function)
about_button.place(x=240, y=0)
excel_home = Button(root, bg='white', bd=0.5, text='H' + "\u0332" + 'ome', padx=20, pady=2, command=home_fuction)
excel_home.place(x=0, y=0)
excel_spliter = Button(root, bg='white', text='S' + "\u0332" + 'plitter', padx=20, pady=4, highlightthickness=0, bd=0.5,
                       command=expliter_fuction)
excel_spliter.place(x=78, y=0)
excel_merge = Button(root, bg='white', text='M' + "\u0332" + 'erge', padx=20, pady=4, highlightthickness=0, bd=0.5,
                     command=merge_fuction)
excel_merge.place(x=160, y=0)
help_text = tk.Text(help_frame)
help_text.tag_configure("center", justify='left')
help_text.insert("1.0", Tips_split)
help_text.tag_add("center", "1.0", "end")

combo_typefile_pop_merge = ttk.Combobox(merge_frame, state="readonly",
                                        value=['Excel file (xlsx)', 'Excel file 97 - 2003 (xls)'])
combo_typefile_pop_merge.current(0)
combo_typefile_pop_merge.place(x=75, y=120)
typefile_info = Label(merge_frame, text='Output\n     File type : ', bg='#a6a6a5', borderwidth=2)
typefile_info.place(x=0, y=105)

output_name_entry = tk.Entry(merge_frame, bd=3, width=27)
output_name_entry_info = Label(merge_frame, text='Output name : ', bg='#a6a6a5', borderwidth=2)
output_name_entry.place(x=100, y=80)
output_name_entry_info.place(x=10, y=80)
text_box_2 = Text(merge_frame, width=57, height=5, borderwidth=2)
text_box_2.place(x=10, y=310, bordermode=INSIDE)
text_box_2.bind("<Key>", lambda a: "break")
scrollbar_2 = Scrollbar(merge_frame)
text_box_2.config(yscrollcommand=scrollbar_2.set)
scrollbar_2.config(command=text_box_2.yview)
scrollbar_2.place(x=475, y=314, relheight=0.2)
showtips_merge = Button(merge_frame, bg='white', text='Help!', padx=40, pady=6, command=help_frame_merge_fuction)
start_merge = Button(merge_frame, bg='white', text='Merge', padx=40, pady=6, command=merge_files_excels)
browse_merge = Button(merge_frame, bg='white', text='Browse Directory ...', padx=25, pady=8,
                      command=upload_action_merge_function)
clear_button_merge = Button(merge_frame, bg='white', text='Clear', padx=10, pady=8, command=clear_button_merge)
clear_button_merge.place(x=225, y=110)
browse_merge_output = Button(merge_frame, bg='white', text='Browse Output Directory ...', padx=25, pady=8,
                             command=directory_choosing_merge_output_function)
browse_merge.place(x=10, y=20)
showtips_merge.place(x=10, y=255)
start_merge.place(x=10, y=210)
browse_merge_output.place(x=10, y=160)

help_text_merge = tk.Text(help_frame_merge)
help_text_merge.tag_configure("center", justify='left')
help_text_merge.insert("1.0", Tips_merge)
help_text_merge.tag_add("center", "1.0", "end")

select_info = Label(frame, text='Header 1', bg='#a6a6a5', borderwidth=2)
select_info_2 = Label(frame, text='Header 2', bg='#a6a6a5', borderwidth=2)
select_info_3 = Label(frame, text='Header 3', bg='#a6a6a5', borderwidth=2)
select_info.place(x=20, y=140)
select_info_2.place(x=20, y=170)
select_info_3.place(x=20, y=200)

directory = ''

tk.Tk.report_callback_exception = show_error

root.mainloop()
